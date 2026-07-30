import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import copy
import io

# 1. 웹페이지 기본 설정
st.set_page_config(page_title="AI스캐너 정산 자동화 시스템", page_icon="🤖", layout="centered")

st.title("🤖 AI스캐너 월별 정산 자동화 시스템")
st.markdown("당월 정산 데이터, 마스터 정보, 전월 정산서 엑셀을 올려주시면 대리점 셀병합 및 매장명 너비가 맞춰진 정산서가 완성됩니다.")

st.divider()

# 2. 최근 3개월 연월 자동 계산 함수
def get_recent_3months(yymm_str):
    yy = int("20" + yymm_str[:2])
    mm = int(yymm_str[2:])
    
    months = []
    for i in range(2, -1, -1):
        m = mm - i
        y = yy
        if m <= 0:
            m += 12
            y -= 1
        months.append(f"{str(y)[2:]}{m:02d}")
    return months

# 안전한 숫자 변환 함수
def safe_float(val, default=0.0):
    try:
        if pd.isna(val): return default
        return float(val)
    except (ValueError, TypeError):
        return default

# 3. 정산 대상 월 입력 섹션
st.subheader("1️⃣ 정산 대상 월 입력")
target_yymm = st.text_input("정산 연월 4자리를 입력하세요 (예: 7월 정산 -> 2607, 8월 정산 -> 2608)", value="2607")

if target_yymm and len(target_yymm) == 4 and target_yymm.isdigit():
    m1, m2, m3 = get_recent_3months(target_yymm)
    st.info(f"📅 **정산 적용 월:** `{m3}` | **최근 3개월 CMS 열 헤더:** `{m1}`, `{m2}`, `{m3}`")
else:
    st.error("⚠️ 정산 연월을 숫자 4자리(예: 2607)로 바르게 입력해 주세요.")

st.divider()

# 4. 파일 업로드 섹션
st.subheader("2️⃣ 엑셀 파일 업로드")

col1, col2 = st.columns(2)

with col1:
    file_cur = st.file_uploader("1) 당월 정산 데이터 (`260728_AI스캐너_정산_...`)", type=["xlsx", "xls"])
    file_master = st.file_uploader("2) 마스터 정산 정보 (`▸AI스캐너_정산정보.xlsx`)", type=["xlsx", "xls"])

with col2:
    file_tpl = st.file_uploader("3) 전월 정산서 템플릿 (`AI스캐너_정산내역_2606_수정.xlsx`)", type=["xlsx", "xls"])

st.divider()

# 5. 정산 처리 및 양식 데이터 자동 채우기
st.subheader("3️⃣ 정산 실행 및 결과 다운로드")

if st.button("🚀 정산 내역서 자동 생성 시작", type="primary", use_container_width=True):
    if not file_cur or not file_master or not file_tpl:
        st.warning("⚠️ 엑셀 파일 3개를 모두 업로드한 후 실행해 주세요.")
    else:
        with st.spinner("⏳ 대리점 셀 병합 및 매장명 너비를 맞추어 엑셀 정산서를 생성하는 중입니다..."):
            try:
                # 1) 전월 정산 내역서 템플릿 파싱
                wb_tpl = openpyxl.load_workbook(file_tpl, data_only=False)
                ws_prev = wb_tpl.worksheets[0]
                
                # 이전달 CMS 이월 상태값 연동 맵 생성
                prev_history_map = {}
                for r in range(7, ws_prev.max_row + 1):
                    store_val = ws_prev.cell(r, 3).value
                    if store_val and str(store_val).strip() != '합계':
                        store_name = str(store_val).strip()
                        k_val = ws_prev.cell(r, 11).value # 전월 K열
                        l_val = ws_prev.cell(r, 12).value # 전월 L열
                        prev_history_map[store_name] = {
                            'col_j_new': k_val if k_val is not None else '',
                            'col_k_new': l_val if l_val is not None else ''
                        }

                # 시트명 및 헤더 연월 업데이트
                ws_prev.title = m3
                ws_prev['C2'] = f"{m3} 대리점 정산(VAT포함)"
                ws_prev['J5'] = f"{m1}(VAT포함)"
                ws_prev['K5'] = f"{m2}(VAT포함)"
                ws_prev['L5'] = f"{m3}(VAT포함)"
                ws_prev['N5'] = f"{m3}(VAT포함)"
                
                # 두 번째 시트(Sheet1 요약표) VLOOKUP 수식 업데이트
                if len(wb_tpl.worksheets) > 1:
                    ws2 = wb_tpl.worksheets[1]
                    ws2['C2'] = f"{m3} 대리점 정산(VAT포함)"
                    for r_idx in range(3, 30):
                        ws2[f'C{r_idx}'] = f"=VLOOKUP(B:B,'{m3}'!B:N,13,0)"

                # 2) 마스터 정산정보 파싱 (CMS 시트)
                df_master_cms = pd.read_excel(file_master, sheet_name='CMS')
                master_map = {}
                for idx, row in df_master_cms.iterrows():
                    store_name = str(row['Unnamed: 1']).strip() if pd.notna(row['Unnamed: 1']) else ''
                    if store_name and store_name != '엑셀':
                        master_map[store_name] = {
                            'agency': str(row['Unnamed: 6']).strip() if pd.notna(row['Unnamed: 6']) else '',
                            'biz_no': str(row['Unnamed: 5']).strip() if pd.notna(row['Unnamed: 5']) else '',
                            'sales_rep': str(row['Unnamed: 8']).strip() if pd.notna(row['Unnamed: 8']) else ''
                        }

                # 3) 당월 정산 데이터 파싱
                df_cur = pd.read_excel(file_cur, sheet_name=0)
                
                store_records = []
                unmapped_stores = []

                for i in range(4, len(df_cur)):
                    row = df_cur.iloc[i]
                    store_name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                    status = str(row.iloc[16]).strip() if pd.notna(row.iloc[16]) else ''

                    if not store_name or store_name == '합계' or store_name == 'nan':
                        continue
                    if status not in ['청구', '미청구']:
                        continue

                    qty_val = row.iloc[15]
                    qty = int(safe_float(qty_val, 1)) if safe_float(qty_val, 1) > 0 else 1
                    
                    cost_ex = safe_float(row.iloc[18], 0.0)
                    pb_ex = safe_float(row.iloc[19], 0.0)
                    
                    cost_inc = round(cost_ex * 1.1)
                    pb_inc = round(pb_ex * 1.1)

                    note = ''
                    if isinstance(row.iloc[18], str) and not str(row.iloc[18]).replace('.','',1).isdigit():
                        note = str(row.iloc[18]).strip()
                    elif isinstance(row.iloc[19], str) and not str(row.iloc[19]).replace('.','',1).isdigit():
                        note = str(row.iloc[19]).strip()

                    master_info = master_map.get(store_name, {})
                    agency = str(row.iloc[11]).strip() if pd.notna(row.iloc[11]) else master_info.get('agency', '미매핑 대리점')
                    biz_no = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else master_info.get('biz_no', '')
                    sales_rep = str(row.iloc[14]).strip() if pd.notna(row.iloc[14]) else master_info.get('sales_rep', '')

                    if not master_info and agency == '미매핑 대리점':
                        unmapped_stores.append(store_name)

                    prev_hist = prev_history_map.get(store_name, {})
                    val_j = prev_hist.get('col_j_new', cost_inc)
                    val_k = prev_hist.get('col_k_new', cost_inc)

                    store_records.append({
                        'agency': agency,
                        'store_name': store_name,
                        'biz_no': biz_no,
                        'sales_rep': sales_rep,
                        'qty': qty,
                        'cost_inc': cost_inc,
                        'pb_inc': pb_inc,
                        'note': note,
                        'val_j': val_j,
                        'val_k': val_k,
                        'val_l': cost_inc
                    })

                df_stores = pd.DataFrame(store_records)
                df_stores.sort_values(by=['agency', 'store_name'], inplace=True)

                # 4) 템플릿 서식을 유지하며 데이터 작성 및 셀 병합
                start_row = 7
                records = df_stores.to_dict('records')
                total_records = len(records)
                end_data_row = start_row + total_records - 1

                # 기준 서식복사 스타일
                template_sample_row = 7
                sample_styles = {}
                for col_i in range(2, 15):
                    sample_cell = ws_prev.cell(template_sample_row, col_i)
                    sample_styles[col_i] = {
                        'font': copy.copy(sample_cell.font),
                        'border': copy.copy(sample_cell.border),
                        'fill': copy.copy(sample_cell.fill),
                        'alignment': copy.copy(sample_cell.alignment),
                        'number_format': sample_cell.number_format
                    }

                agency_rows_map = {}

                max_store_len = 13
                for idx, item in enumerate(records):
                    r = start_row + idx
                    agency = item['agency']

                    if len(item['store_name']) > max_store_len:
                        max_store_len = len(item['store_name'])

                    if agency not in agency_rows_map:
                        agency_rows_map[agency] = []
                    agency_rows_map[agency].append(r)

                    ws_prev[f'B{r}'] = agency
                    ws_prev[f'C{r}'] = item['store_name']
                    ws_prev[f'D{r}'] = item['biz_no']
                    ws_prev[f'E{r}'] = item['sales_rep']
                    ws_prev[f'F{r}'] = item['qty']
                    ws_prev[f'G{r}'] = item['cost_inc']
                    ws_prev[f'H{r}'] = item['pb_inc']
                    ws_prev[f'I{r}'] = item['note'] if item['note'] else None
                    
                    ws_prev[f'J{r}'] = item['val_j']
                    ws_prev[f'K{r}'] = item['val_k']
                    ws_prev[f'L{r}'] = item['val_l']
                    ws_prev[f'M{r}'] = item['pb_inc']

                    # 테두리 및 서식 적용
                    for col_idx in range(2, 15):
                        cell = ws_prev.cell(row=r, column=col_idx)
                        st_dict = sample_styles[col_idx]
                        cell.font = copy.copy(st_dict['font'])
                        cell.border = copy.copy(st_dict['border'])
                        cell.fill = copy.copy(st_dict['fill'])
                        cell.alignment = copy.copy(st_dict['alignment'])
                        if col_idx in [6, 7, 8, 10, 11, 12, 13, 14] and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'

                # ✨ 대리점별 셀 병합 (B열 대리점명, N열 대리점 정산금)
                for agency, row_list in agency_rows_map.items():
                    first_r = row_list[0]
                    last_r = row_list[-1]
                    
                    # 해당 대리점 정산 총액 계산
                    total_pb = sum(records[r - start_row]['pb_inc'] for r in row_list)
                    ws_prev[f'N{first_r}'] = total_pb

                    # 매장이 2개 이상이면 세로 셀 병합 진행
                    if len(row_list) > 1:
                        ws_prev.merge_cells(start_row=first_r, start_column=2, end_row=last_r, end_column=2) # B열 병합
                        ws_prev.merge_cells(start_row=first_r, start_column=14, end_row=last_r, end_column=14) # N열 병합
                        
                        # 병합된 셀 세로 중앙 정렬 지정
                        ws_prev.cell(first_r, 2).alignment = Alignment(horizontal='center', vertical='center')
                        ws_prev.cell(first_r, 14).alignment = Alignment(horizontal='right', vertical='center')

                # ✨ 컬럼 너비 보정 (C열 매장명 너비 자동 확장)
                col_widths = {
                    'A': 3.5, 
                    'B': 22.0, 
                    'C': max(30.0, max_store_len * 1.8), # 매장명 너비 안 잘리게 확장
                    'D': 13.0, 
                    'E': 9.0, 
                    'F': 12.0, 
                    'G': 13.0, 
                    'H': 13.0, 
                    'I': 32.0, 
                    'J': 13.0, 
                    'K': 13.0, 
                    'L': 13.0, 
                    'M': 13.0, 
                    'N': 13.0
                }
                for col_letter, width in col_widths.items():
                    ws_prev.column_dimensions[col_letter].width = width

                # 5) 맨 아래 합계(SUM) 행 작성
                sum_row = end_data_row + 1
                
                ws_prev[f'B{sum_row}'] = "합계"
                ws_prev[f'F{sum_row}'] = f"=SUM(F7:F{end_data_row})"
                ws_prev[f'G{sum_row}'] = f"=SUM(G7:G{end_data_row})"
                ws_prev[f'H{sum_row}'] = f"=SUM(H7:H{end_data_row})"
                ws_prev[f'J{sum_row}'] = f"=SUM(J7:J{end_data_row})"
                ws_prev[f'K{sum_row}'] = f"=SUM(K7:K{end_data_row})"
                ws_prev[f'L{sum_row}'] = f"=SUM(L7:L{end_data_row})"
                ws_prev[f'M{sum_row}'] = f"=SUM(M7:M{end_data_row})"
                ws_prev[f'N{sum_row}'] = f"=SUM(N7:N{end_data_row})"

                sum_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
                thin_border = Border(left=Side(style='thin', color='B0C4DE'), right=Side(style='thin', color='B0C4DE'), top=Side(style='thin', color='B0C4DE'), bottom=Side(style='thin', color='B0C4DE'))
                
                for col_idx in range(2, 15):
                    cell = ws_prev.cell(row=sum_row, column=col_idx)
                    cell.border = thin_border
                    cell.fill = sum_fill
                    cell.font = Font(name='맑은 고딕', size=10, bold=True)
                    if col_idx in [6, 7, 8, 10, 11, 12, 13, 14]:
                        cell.number_format = '#,##0'

                # 6) 메모리 버퍼 저장 및 다운로드
                output_buffer = io.BytesIO()
                wb_tpl.save(output_buffer)
                output_buffer.seek(0)
                
                st.success(f"🎉 `{m3}` 정산 내역서 생성이 완료되었습니다! (동일 대리점 셀병합 & 매장명 너비 보정 완료)")
                
                if unmapped_stores:
                    st.warning(f"⚠️ 당월 데이터 미매핑 매장 ({len(unmapped_stores)}건): {', '.join(unmapped_stores)}")

                st.download_button(
                    label=f"📥 AI스캐너_정산내역_{m3}.xlsx 다운로드",
                    data=output_buffer,
                    file_name=f"AI스캐너_정산내역_{m3}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"❌ 정산 작업 중 오류가 발생했습니다: {e}")
